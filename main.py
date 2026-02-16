import os
import time
import datetime

# import dearpygui
# import dearpygui.dearpygui as dpg
import openpyxl
from openpyxl import Workbook

#qtpython
import sys
import random
from PySide6 import QtCore, QtWidgets, QtGui


wb = Workbook()
ws = wb.active
operating_system = os.name
triggered = False
#strings to look for
#variable_name_start = '<BR><H2><A NAME="'
#variable_name_end = '">'
#test1='					<TABLE>'
#other variables  for loops
#x=0
#print(operating_system) #debug

# comp_file = open("numarul.html","r")
# #print(comp_file.read()) #debug
# lines = len(comp_file.readlines())
# #print(lines) #debug
# comp_file.close()


# comp_file = open("numarul.txt","r")
# variables = comp_file.readlines()
# print(variables[0:232])
# print("second print " + variables[30])
#
#
# time.sleep(6)
# comp_file.close()
# print("Its done")
def main_function(file_path):
    #global file_path
    global new_current_time
    global triggered
    triggered = True
    x = 0
    y = 0
    z = 1
    w = 2
    #shit checking
    comp_file = open(file_path[0], "r")
    # print(comp_file.read()) #debug
    lines = len(comp_file.readlines())
    # print(lines) #debug
    comp_file.close()

    variable_name_start = '<BR><H2><A NAME="'
    variable_list = []
    #print(lines)
    comp_file = open(file_path[0], "r")
    variables = comp_file.readlines()
    #print(variables[0:232])
    #print("second print " + variables[30])
    while x < lines:
        #print(str(x))
        #print(variables[x])
        if "H2><A NAME" in variables[x]:
            #print(x)
            #print(variables[x])
            #print(variables[x].split('"'))
            variable_split = variables[x].split('"')
            #print("god  dammit andy  " +variable_split[1])

            #am adaugat numele variabilei in variable list
            variable_list.append(variable_split[1])
            #print(variable_list)
        #extracting the variable values found in model and in  cdfx file
        #the first if is succesful as long as the text is prezent in the line, but it is not the only thing we need
        #there are variables that are html formatted different and the rule below does not apply and will skip a lot of
        #variables
        if "<TD width=200 nowrap><CODE><font color=" in variables[x]:
            #print(variables[x])
            values_model_and_cdfx = variables[x].split('>')
            values_model_and_cdfx_second_split=values_model_and_cdfx[3].split('<')
            #ADDED THIS IN ORDER REMOVE THE UNITS IN THE MOEL COLUMN
            values_model_and_cdfx_second_split = values_model_and_cdfx[3].split(' ')
            #print(values_model_and_cdfx_second_split[0])

            variable_list.append(values_model_and_cdfx_second_split[0])
        if "<TD VALIGN=top ALIGN=left width=480><font color=" in variables[x]:
            if "Value" in variables[x]:
                #print(variables[x])
                variables_odd_model_and_cdfx = variables[x].split('>')
                #print(variables_odd_model_and_cdfx)
                variables_odd_model_and_cdfx_second_split = variables_odd_model_and_cdfx[2].split('<')
                #print(variables_odd_model_and_cdfx_second_split[0])
                variable_list.append(variables_odd_model_and_cdfx_second_split[0])
        x=x+1
    #everything the software needs to  create the final document
    #print(str(len(variable_list)))
    variable_quantity = len(variable_list) #helps size the number of loops
    # THE NAMING OF THE FILE
    current_time = str(datetime.datetime.now()) #the name of the file
    new_current_time = current_time.replace(':','_') + ".txt"
    #opening the  final .txt file
    final_file=open(new_current_time,"x")
    time.sleep(1)
    #loop to write all the information in the .txt file
    while y<variable_quantity:
        final_file.write(str(variable_list[y]) + '*' + str(variable_list[z]) + '*' + str(variable_list[w]) + '*' + "\n")
        y=y+3
        z=z+3
        w=w+3
    final_file.close()
    #print("Model and cdfx table file created")
    print("File name is " + new_current_time)
    #print("making the file path...")
    new_file_path_maker(file_path[0])


    # with open(new_file_path) as new_file:
    #     print("wait 5 seconds")
    #     time.sleep(5)
    #     print(new_file.readline(44))
    #     new_file_lines = len(new_file.readlines())
    #     print(new_file_lines)



    #print(str(new_file.readline()) +'test')

    #print(new_file_lines)
    # while excel_line<new_file_lines:
    #     #print(new_file[])
    #     print(excel_line)
    #     print(new_file.readline(5))
    #     excel_line = excel_line + 1
    # time.sleep(10)
    #
    # file.close()
    # #creates the excel with correct name an shiet
    # excel_file_name = new_current_time
    # time.sleep(2)
    # wb.save(excel_file_name + ".xlsx")
    # print("everything done, excel created")

    #print(new_file_path[-1])
    #new_file_path_2 = new_file_path[-1].replace('numarul.html',new_current_time)
    #print(new_file_path_2)
#main_function()   #to be uncommented when we finish the GUI vibes

def new_file_path_maker(file_path):
    global new_file_path
    print('this is how it starts:' + str(file_path))
    #THIS CODE MAKES THE NEW FILEPATH TOWARDS THE NEW DOCUMENT THAT WAS JUST CREATED
    new_file_path = file_path.split("//")
    print('wtf:'+ str(new_file_path))
    new_file_path.remove(new_file_path[-1]) #TO SWAP THIS ONE WITH THE ACTUAL NAME OF THE FILE OR MAKE IT new_file_path[-1], not sure
    new_file_path.append(new_current_time)
    new_file_path = '\\'.join(new_file_path)
    print(new_file_path)
    #print('Filepath of the newly created file: ' + str(new_file_path))
    print("making the excel file path..." + new_file_path)

    #make_excel_compare(new_file_path)

def make_excel_compare():
    global excel_file_name
    print('excel file path:'+new_file_path)
    time.sleep(4)
    file_line = 0
    excel_line = 2
    #OPEN THE NEW FILE FROM THE NEW PATH AND  CHECK THE VIBES
    #new_file = open(new_file_path, 'r')
    file = open(new_file_path, "r")
    lines = file.readlines()
    number_of_lines = len(lines)
    print("number of lines: " + str(number_of_lines))
    ws['A1'] = 'Variable Name'
    ws['B1'] = 'Model'
    ws['C1'] = 'CDFX'
    ws['D1'] = 'LG'
    global bitch
    global a_line
    global b_line
    global c_line
    while True:
        bitch = lines[file_line].split('*')
        a_line = ('A' + str(excel_line))
        b_line = ('B' + str(excel_line))
        c_line = ('C' + str(excel_line))
        #print(lines[0].split('*'))
        ws[a_line] = bitch[0]
        ws[b_line] = bitch[1]
        ws[c_line] = bitch[2]
        #equality_checker() #to come back tot this
        file_line = file_line + 1
        excel_line = excel_line + 1
        if excel_line == number_of_lines+2:
            break

    print("closing the file")
    file.close()
    #creates the excel with correct name an shiet
    excel_file_name = new_current_time
    excel_file_name = excel_file_name.split('.')
    excel_file_name = (excel_file_name[0] + '.'+excel_file_name[1] + ".xlsx")

    time.sleep(2)
    wb.save(excel_file_name)
    #wb.close(excel_file_name)
    print("everything done, excel created")

#useless atm as well
def equality_checker():
    value_model = 0
    #hardcode how to handle true and false
    if 'false' in bitch[1]:
        value_model = 0
    else:
        pass

    if 'true' in bitch[1]:
        value_model = 1
    else:
        pass

    if str(value_model) == bitch[2]:
        ws[a_line].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        ws[b_line].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    else:
        pass

#useless
def LG_checker():
    #this function uses E and F colums in order to
    # if A1 == EX
    #   D1 = FX
    ws['D15'] = 'LG'



def excel_path_maker(sender, app_data):
    #print("path made!")
    global excel_path
    #print(app_data)
    #print(new_current_time)
    #excel_path = (app_data['selections'][new_current_time])
    if triggered == True:
        excel_path = app_data['selections'][excel_file_name]
        print(excel_path)
    else:
        print("you must run the previous step")

def load_and_check_LG(file_path):
    variable_line = 2 #for the name of the variable #A
    LG_variable_name_line = 2 #name of the variable in the LG column #E
    LG_variable_value_line = 2  #value of the LG variable must be on the same line as the name #F
    variable_placement = 2 #where to put the value when you find the match in the LG #D
    print(file_path)
    #opening the excel
    wb=openpyxl.load_workbook(file_path[0])
    ws = wb.active
    #MAKING THE NUMBERS FOR THE EXCEL CASETTE
    # A_spot = ('A' + str(variable_line))
    # E_spot = ('E' + str(LG_variable_name_line))
    # F_spot = ('F' + str(LG_variable_value_line))
    # D_spot = ('D' + str(variable_placement))
    #VALUE OF THE CASSETES
    # A_column = ws[A_spot].value
    # E_column = ws[E_spot].value
    # F_column = ws[F_spot].value
    # D_column = ws[D_spot].value
    #print("A Column: " + A_column)
    #print("E Column: " + E_column)
    #ws[D_spot] = 'bitch'

    while True:
        A_spot = ('A' + str(variable_line))
        E_spot = ('E' + str(LG_variable_name_line))
        F_spot = ('F' + str(LG_variable_value_line))
        D_spot = ('D' + str(variable_placement))
        # VALUE OF THE CASSETES
        A_column = ws[A_spot].value
        E_column = ws[E_spot].value
        F_column = ws[F_spot].value
        D_column = ws[D_spot].value
        if A_column == None:
            break
        # if E_column == None:
        #     break
        #daca variabilele au acelasi nume, trecem la urmatoarea variabila
        if A_column == E_column:
            ws[D_spot] = F_column
            #time.sleep(1)
            variable_line = variable_line +1
            variable_placement=variable_placement+1
            LG_variable_value_line = 2
            LG_variable_name_line = 2
            print(str(variable_line))
        else:   #daca nu au acelasi nume, mergem mai departe
            LG_variable_name_line = LG_variable_name_line+1
            LG_variable_value_line = LG_variable_value_line+1
            print("LG varriable name line: " + str(LG_variable_name_line))
            print("LG_variable_value_line: " + str(LG_variable_value_line))
            if E_column == None:
                print(E_column)
                variable_line = variable_line + 1
                variable_placement = variable_placement + 1
                LG_variable_value_line = 2
                LG_variable_name_line = 2



        wb.save(file_path[0])
    print("Compare finished")

#IMPLEMENTING A LITTLE JUICY THREADING TO MAKE IT MOVE FASTER

# def callback(sender, app_data):
#     print('OK was clicked.')
#     print(app_data['selections']['numarul.html'])
#     main_function(app_data['selections']['numarul.html'])
#
# def cancel_callback(sender, app_data):
#     print('Cancel was clicked.')
#     print("Sender: ", sender)
#     print("App Data: ", app_data)
#
# dpg.create_context()
#
# with dpg.file_dialog(directory_selector=False, show=False, callback=callback, file_count=3, tag="file_dialog_tag", width=700 ,height=400):
#     dpg.add_file_extension("", color=(255, 150, 150, 255))
#     dpg.add_file_extension(".*")
#     dpg.add_file_extension(".cpp", color=(255, 255, 0, 255))
#     dpg.add_file_extension(".h", color=(255, 0, 255, 255))
#     dpg.add_file_extension(".py", color=(0, 255, 0, 255))
#
#     dpg.add_button(label="fancy file dialog")
#     with dpg.child_window(width=100):
#         dpg.add_selectable(label="bookmark 1")
#         dpg.add_selectable(label="bookmark 2")
#         dpg.add_selectable(label="bookmark 3")
#
# with dpg.file_dialog(directory_selector=False, show=False, callback=excel_path_maker, file_count=3, tag="file_dialog_tag2", width=700 ,height=400):
#     dpg.add_file_extension("", color=(255, 150, 150, 255))
#     dpg.add_file_extension(".*")
#     dpg.add_file_extension(".cpp", color=(255, 255, 0, 255))
#     dpg.add_file_extension(".h", color=(255, 0, 255, 255))
#     dpg.add_file_extension(".py", color=(0, 255, 0, 255))
#
#     dpg.add_button(label="fancy file dialog")
#     with dpg.child_window(width=100):
#         dpg.add_selectable(label="bookmark 1")
#         dpg.add_selectable(label="bookmark 2")
#         dpg.add_selectable(label="bookmark 3")
#
# with dpg.window(tag="Florin Miracle"):
#     dpg.add_text("Hello")
#     dpg.add_text("1. First step is to click the 'Browse HTML' button below and choosethe HTML compare file")
#     dpg.add_button(label="Browse HTML", callback=lambda: dpg.show_item("file_dialog_tag"))
#     dpg.add_text("2. After you select the HTML file, press the START button below in order to make the excel file")
#     dpg.add_button(label="START", callback=make_excel_compare)
#     dpg.add_text("3. After pressing start an EXCEL document will be created named after the current date and hour")
#     dpg.add_text("4. Next step is to find in the LG the sheet with the variables")
#     dpg.add_text("5. Take the column with names of variables from the LG and but it on E2 and the values accordingly on F2")
#     dpg.add_text("6. DO NOT RENAME THE EXCEL, after you add the LG information browse for the excel with the Browse for Excel button")
#     dpg.add_button(label="Browse for Excel", callback=lambda: dpg.show_item("file_dialog_tag2"))
#     dpg.add_text("7. After getting the excel ready, press the FINISH HIM button and it will search for the variable in the LG column and add its value in the column D")
#     dpg.add_button(label="FINISH HIM", callback=load_and_check_LG)
#
#     #dpg.add_button(label="Directory Selector", callback=lambda: dpg.show_item("file_dialog_id"))
#     #dpg.add_button(label="File Selector", callback=lambda: dpg.show_item("file_dialog_tag"))
#
#
#
#
# dpg.create_viewport(title='Florin Miracle', width=1200, height=500)
# dpg.setup_dearpygui()
# dpg.show_viewport()
# dpg.set_primary_window("Florin Miracle", True)
# dpg.start_dearpygui()
# dpg.destroy_context()




class MyWidget(QtWidgets.QWidget):
    def __init__(self):
        super().__init__()
        #declare list
        self.hello = ["Hallo Welt", "Hei maailma", "Hola Mundo", "Привет мир"]
        #declare widgets
        #self.button = QtWidgets.QPushButton("Click me!")
        self.text = QtWidgets.QLabel("Hello Bajetzii", alignment=QtCore.Qt.AlignCenter)
        self.text2 = QtWidgets.QLabel("Use the Browse for HTML button to look for the HTML compare file, it will create the classic .txt", alignment=QtCore.Qt.AlignCenter)
        self.text3 = QtWidgets.QLabel("Press the Make the excel button to create an excel file based on the .txt file created earlier", alignment=QtCore.Qt.AlignCenter)
        self.text4 = QtWidgets.QLabel("Open the created excel and add the variable and value columns found in LG, variables starting E2 and values starting F2", alignment=QtCore.Qt.AlignCenter)
        self.button2 = QtWidgets.QPushButton("Browse for HTML")
        self.button3 = QtWidgets.QPushButton("Make the excel")
        self.button4 = QtWidgets.QPushButton("Browse for excel")
        # file selection

        self.button2.clicked.connect(self.open_file_dialog_html)
        self.button3.clicked.connect(make_excel_compare)
        self.button4.clicked.connect(self.open_file_dialog_excel)
        #self.button = QtWidgets.QFileDialog
        #create the shit
        self.layout = QtWidgets.QVBoxLayout(self)
        self.layout.addWidget(self.text)
        self.layout.addWidget(self.text2)
        self.layout.addWidget(self.text3)
        self.layout.addWidget(self.text4)
        #self.layout.addWidget(self.button)
        self.layout.addWidget(self.button2)
        self.layout.addWidget(self.button3)
        self.layout.addWidget(self.button4)
        #conenct with function
        #self.button.clicked.connect(self.magic)

    def open_file_dialog_html(self):
        #global file_path
        dialog = QtWidgets.QFileDialog(self)
        #dialog.setDirectory(r'C:\images')
        dialog.setFileMode(QtWidgets.QFileDialog.FileMode.ExistingFiles)
        dialog.setNameFilter("Files (*.xlsx *.html *.txt *.HTM)")
        dialog.setViewMode(QtWidgets.QFileDialog.ViewMode.List)
        if dialog.exec():
            file_path = dialog.selectedFiles()
            print(file_path)
            main_function(file_path)

            # if filenames:
            #     self.file_list.addItems([str(Path(filename)) for filename in filenames])

    def open_file_dialog_excel(self):
        global file_path
        dialog = QtWidgets.QFileDialog(self)
        #dialog.setDirectory(r'C:\images')
        dialog.setFileMode(QtWidgets.QFileDialog.FileMode.ExistingFiles)
        dialog.setNameFilter("Files (*.HTM *.xlsx *.html *.txt)")
        dialog.setViewMode(QtWidgets.QFileDialog.ViewMode.List)
        if dialog.exec():
            file_path = dialog.selectedFiles()
            #print(file_path)
            load_and_check_LG(file_path)

    @QtCore.Slot()
    def magic(self):
        self.text.setText(random.choice(self.hello))

if __name__ == "__main__":
    app = QtWidgets.QApplication([])

    widget = MyWidget()
    widget.resize(800, 600)
    widget.show()

    sys.exit(app.exec())


